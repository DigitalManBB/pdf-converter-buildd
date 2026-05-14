import os
import logging
from pathlib import Path
from io import BytesIO

import pandas as pd
import pdfplumber
from pdf2image import convert_from_path
from img2table.document import Image as Img2TableImage
from img2table.ocr import TesseractOCR
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

import sys
import os

def resource_path(relative_path):
    """Получает абсолютный путь к ресурсу, работает и для dev, и для PyInstaller."""
    try:
        # PyInstaller создает временную папку и хранит путь в _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Указываем программе, где лежат её инструменты
tesseract_exe = resource_path(os.path.join('tesseract_portable', 'tesseract.exe'))
if os.path.exists(tesseract_exe):
    pytesseract.pytesseract.tesseract_cmd = tesseract_exe

poppler_bin = resource_path(os.path.join('poppler_portable', 'Library', 'bin'))
if os.path.exists(poppler_bin):
    os.environ['PATH'] = poppler_bin + ';' + os.environ.get('PATH', '')

OUTPUT_DIR = "output_new"
DPI = 300
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
ocr_engine = TesseractOCR(lang="rus+eng")


# ---------- ШАПКА ----------
def extract_header_text_textpdf(pdf_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[0].extract_text()
            lines = text.split('\n')
            for i, line in enumerate(lines):
                if any(kw in line for kw in ['№ пп', '№', 'Обоснование', 'Наименование', 'Ед. изм.']):
                    return lines[:i]
            return []
    except:
        return []


def extract_header_text_ocr(img):
    try:
        import pytesseract
        text = pytesseract.image_to_string(img, lang='rus')
        return text.strip().split('\n')[:30]
    except:
        return ["[Шапка не распознана (скан)]"]


# ---------- ВСПОМОГАТЕЛЬНАЯ: ВЫРАВНИВАНИЕ ЗАГОЛОВКА ----------
def align_header_to_df(df, header_candidates, default_header):
    """
    Приводит заголовок к количеству колонок DataFrame.
    """
    num_cols = df.shape[1]

    # Если нашли строку с ключевыми словами
    header_row_idx = None
    for idx, row in df.iterrows():
        row_str = ' '.join(str(x) for x in row if x)
        if any(kw in row_str for kw in header_candidates):
            header_row_idx = idx
            break

    if header_row_idx is not None:
        # Используем найденную строку как заголовок
        new_header = df.iloc[header_row_idx].tolist()
        # Удаляем строку заголовка из данных
        df = df.drop(header_row_idx).reset_index(drop=True)
    else:
        new_header = default_header.copy()

    # ВАЖНО: обрезаем или дополняем заголовок до размера df
    if len(new_header) < num_cols:
        new_header.extend([''] * (num_cols - len(new_header)))
    elif len(new_header) > num_cols:
        new_header = new_header[:num_cols]

    df.columns = new_header
    return df


# ---------- СБОР ВСЕХ ТАБЛИЦ (PDFPlumber) ----------
def extract_all_tables_pdfplumber(pdf_path):
    all_tables = []
    keywords = ['№ пп', 'Обоснование', 'Наименование', 'Ед. изм.', 'Кол-во', 'Сметная стоимость']
    default_header = ['№', 'Обоснование', 'Наименование работ', 'Ед. изм.', 'Кол.', 'Сметная стоимость']

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables(table_settings={
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "snap_tolerance": 3,
                    "join_tolerance": 5
                })
                for tbl in tables:
                    if not tbl or len(tbl) < 2:
                        continue

                    df = pd.DataFrame(tbl)
                    df = df.replace(r'\n+', ' ', regex=True)
                    df = df.replace('', None).dropna(how='all').dropna(axis=1, how='all').fillna('')

                    if df.empty:
                        continue

                    # Автоматическое назначение заголовка
                    df = align_header_to_df(df, keywords, default_header)
                    all_tables.append(df)
        return all_tables
    except Exception as e:
        logger.error(f"pdfplumber error: {e}")
        return []


# ---------- СБОР ВСЕХ ТАБЛИЦ (OCR) ----------
def extract_all_tables_ocr(pdf_path):
    all_tables = []
    keywords = ['№ пп', 'Обоснование', 'Наименование', 'Ед. изм.', 'Кол-во', 'Сметная стоимость']
    default_header = ['№', 'Обоснование', 'Наименование работ', 'Ед. изм.', 'Кол.', 'Сметная стоимость']

    try:
        images = convert_from_path(pdf_path, dpi=DPI)
        for img in images:
            buf = BytesIO()
            img.save(buf, format='PNG')
            buf.seek(0)
            doc = Img2TableImage(buf)
            tables = doc.extract_tables(ocr=ocr_engine, implicit_rows=True, borderless_tables=True)
            for table in tables:
                if table.df is not None and not table.df.empty:
                    df = table.df.copy()
                    df = align_header_to_df(df, keywords, default_header)
                    all_tables.append(df)
        return all_tables
    except Exception as e:
        logger.error(f"OCR error: {e}")
        return []


# ---------- MAIN ----------
def main():
    input_dir = Path("input_pdfs")
    output_dir = Path(OUTPUT_DIR)
    output_dir.mkdir(exist_ok=True)

    for pdf_path in input_dir.glob("*.pdf"):
        logger.info(f"Обработка: {pdf_path.name}")
        try:
            # 1. Шапка
            header_lines = extract_header_text_textpdf(str(pdf_path))
            if not header_lines:
                try:
                    imgs = convert_from_path(str(pdf_path), dpi=DPI, first_page=1, last_page=1)
                    if imgs:
                        header_lines = extract_header_text_ocr(imgs[0])
                except:
                    header_lines = ["[Шапка не распознана]"]

            # 2. Собираем все таблицы
            all_tables = extract_all_tables_pdfplumber(str(pdf_path))
            if not all_tables:
                logger.info("  pdfplumber не нашёл таблиц. Запуск OCR...")
                all_tables = extract_all_tables_ocr(str(pdf_path))

            if not all_tables:
                logger.warning(f"  Не найдено ни одной таблицы в {pdf_path.name}")
                continue

            # 3. Создаём Excel (1 лист)
            wb = Workbook()
            ws = wb.active
            ws.title = "Смета"

            # 3.1. Шапка в первую колонку
            row = 1
            for line in header_lines:
                ws.cell(row=row, column=1, value=line)
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
                row += 1
            if header_lines:
                row += 1  # отступ

            # 3.2. Записываем все таблицы подряд
            for idx, df in enumerate(all_tables):
                if idx > 0:
                    row += 1  # разделитель между таблицами

                # Заголовки таблицы
                headers = list(df.columns)
                for col_idx, h in enumerate(headers, 1):
                    ws.cell(row=row, column=col_idx, value=h)
                row += 1

                # Данные таблицы
                for data_row in df.values:
                    for col_idx, val in enumerate(data_row, 1):
                        ws.cell(row=row, column=col_idx, value=val)
                    row += 1

            # 3.3. Настройка ширины колонок
            if all_tables:
                max_cols = max(df.shape[1] for df in all_tables)
                for c in range(1, max_cols + 1):
                    col_letter = get_column_letter(c)
                    if c == 3:
                        ws.column_dimensions[col_letter].width = 50
                    elif c <= 7:
                        ws.column_dimensions[col_letter].width = 15
                    else:
                        ws.column_dimensions[col_letter].width = 20

            # Перенос текста для всех ячеек
            for r in ws.iter_rows(min_row=1):
                for cell in r:
                    if cell.value:
                        cell.alignment = Alignment(wrap_text=True, horizontal='left')

            # Сохраняем
            out_file = output_dir / (pdf_path.stem + ".xlsx")
            wb.save(out_file)
            logger.info(f"  Сохранён: {out_file.name} (таблиц: {len(all_tables)})")

        except Exception as e:
            logger.exception(f"Ошибка при обработке {pdf_path.name}: {e}")

    logger.info("Готово. Все таблицы на одном листе с заголовками.")


if __name__ == "__main__":
    main()
